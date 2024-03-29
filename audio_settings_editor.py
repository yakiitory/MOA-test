import os
import random
import shutil
import sys
import tkinter as tk
from tkinter import filedialog, Toplevel
from excel_export import export_to_excel
from openpyxl import Workbook, load_workbook

class AudioSettingsEditor:
    def __init__(self):
        self.file_path = ""
        self.pm_folder_path = os.path.join(os.getcwd(), "music")  # Get the path to the "music" folder in the current directory
        self.track_paths = {}

        # Look for audio files numbered from 1 to 5 in the "pm" folder with any audio format
        for i in range(1, 6):
            for file_name in os.listdir(self.pm_folder_path):
                # Check if the file starts with the number and has a recognized audio file extension
                if file_name.startswith(str(i)) and any(file_name.lower().endswith(ext) for ext in ['.mp3', '.wav', '.flac']):
                    self.track_paths[i] = os.path.join(self.pm_folder_path, file_name)
                    break

        self.create_initial_screen()

    def create_initial_screen(self):
        self.root = tk.Tk()  # Create the root window
        self.root.withdraw()  # Hide the initial screen
        self.file_path = filedialog.askopenfilename(initialdir="/", title="Select Preference.txt",
                                                    filetypes=[("Text files", "*.txt")])
        if self.file_path:
            self.randomize_filter_values()
            self.open_trial_screen(1)

    # Inside the AudioSettingsEditor class

    def open_trial_screen(self, trial_num):
        trial_screen: Toplevel = tk.Toplevel(self.root)
        trial_screen.title(f"Trial {trial_num} of 5")

        label_title = tk.Label(trial_screen, text="Listening Test Software", font=("Arial", 24, "bold"))
        label_title.pack(pady=20)

        label_subtitle = tk.Label(trial_screen,
                                  text="Adjust the headphone's sound until it matches your preference",
                                  font=("Arial", 16))
        label_subtitle.pack(pady=10)

        frame_buttons = tk.Frame(trial_screen)
        frame_buttons.pack(pady=50)

        eq_buttons = [("Tonality", 1), ("Bass", 3), ("Treble", 4), ("Intensity", 5)]
        for label_text, filter_num in eq_buttons:
            label = tk.Label(frame_buttons, text=label_text, font=("Arial", 16))
            label.grid(row=filter_num - 1, column=0, padx=5, pady=5)
            if label_text == "Tonality":
                up_button = tk.Button(frame_buttons, text="▲", font=("Arial", 16),
                                      command=lambda filter_num=filter_num: (
                                          self.update_tonality_value(True),
                                          self.calculate_filter_6_value()))  # Combine both function calls into one tuple
                down_button = tk.Button(frame_buttons, text="▼", font=("Arial", 16),
                                        command=lambda filter_num=filter_num: (
                                            self.update_tonality_value(False),
                                            self.calculate_filter_6_value()))  # Combine both function calls into one tuple
            else:
                up_button = tk.Button(frame_buttons, text="▲", font=("Arial", 16),
                                      command=lambda filter_num=filter_num: (
                                          self.update_filter_value(filter_num, True),
                                          self.calculate_filter_6_value()))  # Combine both function calls into one tuple
                down_button = tk.Button(frame_buttons, text="▼", font=("Arial", 16),
                                        command=lambda filter_num=filter_num: (
                                            self.update_filter_value(filter_num, False),
                                            self.calculate_filter_6_value()))  # Combine both function calls into one tuple
            up_button.grid(row=filter_num - 1, column=1, padx=5, pady=5)
            down_button.grid(row=filter_num - 1, column=2, padx=5, pady=5)

        play_button = tk.Button(trial_screen, text="Play Audio", font=("Arial", 16),
                                command=lambda trial_num=trial_num: self.play_audio(trial_num))
        play_button.pack(pady=20)

        next_button = tk.Button(trial_screen, text="Next", font=("Arial", 16),
                                command=lambda: self.save_preference(trial_num))
        next_button.pack(pady=20)

        # Add an exit button to close the trial screen
        exit_button = tk.Button(trial_screen, text="Exit", font=("Arial", 16), command=lambda: sys.exit(0))
        exit_button.pack(pady=20)

        # Call calculate_filter_6_value() when opening a trial screen
        self.calculate_filter_6_value()

        self.pages = {}
        self.pages[trial_num] = trial_screen
        try:
            trial_screen.mainloop()  # Ensure the trial screen is displayed
        except KeyboardInterrupt:
            print("Program is being closed...")
            self.root.destroy()

    def create_ending_screen(self):
        ending_screen = tk.Toplevel()
        ending_screen.title("End of Trials")

        label_title = tk.Label(ending_screen, text="Thank you for participating", font=("Arial", 24, "bold"))
        label_title.pack(pady=20)

        # Destroy the root window when ending screen is displayed
        self.root.destroy()

        ending_screen.mainloop()

    def randomize_filter_values(self):
        filter1_value = round(random.uniform(-5, 5) / 0.25) * 0.25
        if filter1_value > 0:
            filter2_value = -filter1_value
        else:
            filter2_value = -filter1_value
        self.set_filter_value(1, filter1_value)
        self.set_filter_value(2, filter2_value)
        for i in range(3, 6):
            filter_value = round(random.uniform(-5, 5) / 0.25) * 0.25
            self.set_filter_value(i, filter_value)

    def set_filter_value(self, filter_num, value):
        if not self.file_path:
            print("Error: Preference.txt not uploaded.")
            return

        lines = self.read_file(self.file_path)
        for i, line in enumerate(lines):
            if f"Filter {filter_num}" in line:
                parts = line.split(" ")
                for j, part in enumerate(parts):
                    if part == "Gain":
                        parts[j + 1] = "{:.2f}".format(value)
                lines[i] = " ".join(parts)

        self.update_file(self.file_path, lines)
        self.print_preference()

    def update_filter_value(self, filter_num, up):
        if not self.file_path:
            print("Error: Preference.txt not uploaded.")
            return

        lines = self.read_file(self.file_path)
        for i, line in enumerate(lines):
            if f"Filter {filter_num}" in line:
                parts = line.split(" ")
                for j, part in enumerate(parts):
                    if part == "Gain":
                        current_db = float(parts[j + 1])
                        parts[j + 1] = "{:.2f}".format(current_db + 0.25 if up else current_db - 0.25)
                        new_db = current_db + 0.25 if up else current_db - 0.25
                        # Check if Bass value became 0 after adjusting Tonality or Filter 3 directly
                        if filter_num == 3 and new_db == 0:
                            # Set both Filter 3 and Filter 6 values to 0
                            self.set_filter_value(3, 0)
                            self.set_filter_value(6, 0)
                            return
                        # Set both Filter 2 and Filter 6 values to 0
                        if filter_num == 2 and new_db == 0:
                            self.set_filter_value(2, 0)
                            self.set_filter_value(6, 0)
                            return
                        lines[i] = " ".join(parts)

        # Update Filter 6 value when Filter 2 is adjusted
        if filter_num == 2:
            self.set_filter_value(6,
                                  0)  # Set value as 0 for now, it will be updated based on the condition in set_filter_value

        self.update_file(self.file_path, lines)
        self.print_preference()

    def update_tonality_value(self, up):
        if not self.file_path:
            print("Error: Preference.txt not uploaded.")
            return

        lines = self.read_file(self.file_path)
        for i, line in enumerate(lines):
            if "Filter 1" in line:
                parts = line.split(" ")
                for j, part in enumerate(parts):
                    if part == "Gain":
                        current_db = float(parts[j + 1])
                        parts[j + 1] = "{:.2f}".format(current_db - 0.25 if up else current_db + 0.25)
                lines[i] = " ".join(parts)
            elif "Filter 2" in line:
                parts = line.split(" ")
                for j, part in enumerate(parts):
                    if part == "Gain":
                        current_db = float(parts[j + 1])
                        parts[j + 1] = "{:.2f}".format(current_db + 0.25 if up else current_db - 0.25)
                lines[i] = " ".join(parts)

        self.update_file(self.file_path, lines)
        self.print_preference()

    def calculate_filter_6_value(self):
        if not self.file_path:
            print("Error: Preference.txt not uploaded.")
            return

        lines = self.read_file(self.file_path)

        # Check the value of filter 3
        filter_3_value = self.get_filter_value(lines, 3)
        if filter_3_value == 0:
            # If filter 3 value is 0, set filter 6 to 0
            self.set_filter_6_value(0)
            return

        # If filter 3 value is not 0, proceed with the next condition
        filter_2_value = self.get_filter_value(lines, 2)
        if filter_2_value == 0:
            # If filter 2 value is 0, set filter 6 to 0
            self.set_filter_6_value(0)
            return

        if filter_2_value != 0:
            # Count the steps of 0.25 dB in filter 2
            steps = abs(filter_2_value) / 0.25
            # Multiply steps by 0.175 dB
            filter_6_value = steps * 0.175
            # Adjust the sign based on the sign of filter 2
            filter_6_value *= 1 if filter_2_value > 0 else -1
            # Write the calculated value to filter 6
            self.set_filter_6_value(filter_6_value)

    def get_filter_value(self, lines, filter_num):
        # Helper method to extract the gain value of a filter from the lines of Preference.txt
        for line in lines:
            if f"Filter {filter_num}" in line:
                parts = line.split(" ")
                for part in parts:
                    if part == "Gain":
                        return float(parts[parts.index(part) + 1])
        return None

    def set_filter_6_value(self, value):
        # Set the value of filter 6 in Preference.txt
        if not self.file_path:
            print("Error: Preference.txt not uploaded.")
            return

        lines = self.read_file(self.file_path)
        for i, line in enumerate(lines):
            if "Filter 6" in line:
                parts = line.split(" ")
                for j, part in enumerate(parts):
                    if part == "Gain":
                        parts[j + 1] = "{:.2f}".format(value)
                lines[i] = " ".join(parts)

        self.update_file(self.file_path, lines)
        self.print_preference()
    def read_file(self, file_path):
        with open(file_path, "r") as file:
            lines = file.readlines()
        return lines

    def update_file(self, file_path, lines):
        with open(file_path, "w") as file:
            file.writelines(lines)
    def play_audio(self, trial_num):
        track_path = self.track_paths.get(trial_num)
        if track_path:
            os.startfile(track_path)
        else:
            print(f"Track path not found for trial {trial_num}.")
    def combine_excel_files(self, directory, combined_file):
            combined_wb = Workbook()
            combined_ws = combined_wb.active
            combined_ws.append(["Trial", "Preference Adjustments", "Value (dB)"])

            for i in range(1, 6):
                excel_file = os.path.join(directory, f"PF{i}.xlsx")
                if os.path.exists(excel_file):
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        row_values = [f"Trial {i}"] + [cell.value for cell in row]
                        combined_ws.append(row_values)

            combined_wb.save(combined_file)
    def save_preference(self, trial_num):
        # Read Preference.txt
        lines = self.read_file(self.file_path)

        # Save a copy of Preference.txt
        preference_dir = "data"
        if not os.path.exists(preference_dir):
            os.makedirs(preference_dir)
        shutil.copy(self.file_path, os.path.join(preference_dir, f"PF{trial_num}.txt"))

        # Get the values for tonality, bass, treble, and intensity from Filter 2, Filter 3, Filter 4, and Filter 5 respectively
        tonality_value = self.get_filter_value(lines, 2)
        if tonality_value < 0:
            tonality_value *= -2  # If negative, double the absolute value
        elif tonality_value > 0:
            tonality_value *= -2  # If positive, make negative
        # If tonality_value is zero, no change is needed
        bass_value = self.get_filter_value(lines, 3)
        treble_value = self.get_filter_value(lines, 4)
        intensity_value = self.get_filter_value(lines, 5)

        # Export lines to Excel
        excel_file = os.path.join("data", f"PF{trial_num}.xlsx")
        export_to_excel(excel_file, lines, tonality_value, bass_value, treble_value, intensity_value)

        # Create combined CSV file when PF5.xlsx is created
        if trial_num == 5:
            csv_file = os.path.join("data", "PF.csv")

        # Close the current trial screen
        self.pages[trial_num].destroy()

        # Open the next trial screen if it's not the last one
        next_trial = trial_num + 1
        if next_trial <= 5:
            self.randomize_filter_values()
            self.open_trial_screen(next_trial)
        # Open the ending screen if it's the last trial
        if trial_num == 5:
            self.root.deiconify()
            self.create_ending_screen()
            combined_excel_file = os.path.join("data", "PF_combined.xlsx")
            self.combine_excel_files("data", combined_excel_file)
    def print_preference(self):
        if self.file_path:
            print("Preference.txt contents:")
            with open(self.file_path, "r") as file:
                for line in file:
                    print(line.strip())
        else:
            print("Error: Preference.txt not uploaded.")

    # Create an instance of AudioSettingsEditor to start the application


app = AudioSettingsEditor()
